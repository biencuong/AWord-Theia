#!/usr/bin/env python3
"""Create ND30-related .xlsx workbooks from a JSON spec.

Usage:
    python create_nd30_workbook.py input.json output.xlsx
"""
from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

THIN = Side(style="thin", color="000000")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
SUBTITLE_FILL = PatternFill("solid", fgColor="EAF2F8")
TITLE_FILL_KEY = "1F4E78"
TYPE_MAP = {
    "so-dang-ky-van-ban-di": [
        "Số, ký hiệu văn bản", "Ngày tháng văn bản", "Tên loại và trích yếu nội dung văn bản",
        "Người ký", "Nơi nhận văn bản", "Đơn vị, người nhận bản lưu", "Số lượng bản",
        "Ngày chuyển", "Ký nhận", "Ghi chú",
    ],
    "so-dang-ky-van-ban-den": [
        "Ngày đến", "Số đến", "Tác giả", "Số, ký hiệu văn bản", "Ngày tháng văn bản",
        "Tên loại và trích yếu nội dung văn bản", "Đơn vị hoặc người nhận", "Ngày chuyển",
        "Ký nhận", "Ghi chú",
    ],
    "so-theo-doi-giai-quyet-van-ban-den": [
        "Số đến", "Tên loại, số, ký hiệu, ngày tháng và tên cơ quan ban hành văn bản",
        "Đơn vị hoặc người nhận", "Thời hạn giải quyết", "Tiến độ giải quyết",
        "Số, ký hiệu văn bản trả lời", "Ghi chú",
    ],
    "muc-luc-ho-so": [
        "Số TT", "Số, ký hiệu hồ sơ", "Tiêu đề hồ sơ", "Thời gian tài liệu",
        "Thời hạn bảo quản", "Số tờ/Số trang", "Ghi chú",
    ],
    "muc-luc-van-ban": [
        "STT", "Số, ký hiệu văn bản", "Ngày tháng năm văn bản", "Tên loại và trích yếu nội dung văn bản",
        "Tác giả văn bản", "Tờ số/Trang số", "Ghi chú",
    ],
    "danh-sach-noi-nhan": [
        "STT", "Tên cơ quan, tổ chức hoặc cá nhân", "Nhóm nơi nhận", "Vai trò nhận",
        "Địa chỉ thư điện tử", "Ghi chú",
    ],
    "phu-luc-bang": None,
}


def load_json(path: Path) -> dict[str, Any]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, dict):
        raise ValueError("JSON root must be an object")
    return data


def apply_cell_style(cell, *, bold=False, fill=None, size=11, align="center", wrap=True):
    color = "FFFFFF" if getattr(fill, 'fgColor', None) and fill.fgColor.rgb and fill.fgColor.rgb[-6:] == TITLE_FILL_KEY else "000000"
    cell.font = Font(name="Times New Roman", size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill:
        cell.fill = fill
    cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def autosize(ws):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            value = str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), min(max(len(value) + 2, 10), 42))
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_title_block(ws, title: str, subtitle: str = "") -> int:
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = title
    apply_cell_style(c, bold=True, fill=TITLE_FILL, size=13)
    row = 2
    if subtitle:
        ws.merge_cells("A2:G2")
        c2 = ws["A2"]
        c2.value = subtitle
        apply_cell_style(c2, bold=True, fill=SUBTITLE_FILL, size=11)
        row = 3
    return row


def build_single_sheet(ws, spec: dict[str, Any]):
    wb_type = spec.get("workbook_type", "phu-luc-bang")
    title = spec.get("title") or wb_type.replace("-", " ").upper()
    subtitle = spec.get("subtitle", "")
    start_row = write_title_block(ws, title, subtitle)
    columns = spec.get("columns")
    if wb_type != "phu-luc-bang":
        columns = columns or TYPE_MAP[wb_type]
    if not columns:
        raise ValueError("columns are required for phu-luc-bang")
    header_row = start_row
    for idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=idx, value=name)
        apply_cell_style(cell, bold=True, fill=HEADER_FILL, size=11)
    rows = spec.get("rows", [])
    for r_offset, row_data in enumerate(rows, start=1):
        row_idx = header_row + r_offset
        if isinstance(row_data, dict):
            ordered = [row_data.get(key, "") for key in columns]
        else:
            ordered = list(row_data)
        for c_idx, value in enumerate(ordered, start=1):
            cell = ws.cell(row=row_idx, column=c_idx, value=value)
            apply_cell_style(cell, size=11, align="left" if c_idx not in (1, 2, 4, 7, 8) else "center")
    ws.freeze_panes = f"A{header_row+1}"
    ws.sheet_view.showGridLines = False
    autosize(ws)


def create_workbook(spec: dict[str, Any], out_path: Path):
    wb = Workbook()
    default = wb.active
    sheets = spec.get("sheets")
    if sheets:
        default.title = sheets[0].get("sheet_name", "Sheet1")
        build_single_sheet(default, sheets[0])
        for extra in sheets[1:]:
            ws = wb.create_sheet(title=extra.get("sheet_name", "Sheet"))
            build_single_sheet(ws, extra)
    else:
        default.title = spec.get("sheet_name", "Du-lieu")
        build_single_sheet(default, spec)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main(argv: list[str]) -> int:
    if len(argv) != 3:
        print("Usage: python create_nd30_workbook.py input.json output.xlsx")
        return 1
    input_path = Path(argv[1])
    output_path = Path(argv[2])
    if not input_path.exists():
        print(f"Input file not found: {input_path}")
        return 1
    spec = load_json(input_path)
    wb_type = spec.get("workbook_type")
    if not wb_type and not spec.get("sheets"):
        print("workbook_type is required unless sheets are provided")
        return 1
    if wb_type and wb_type not in TYPE_MAP:
        print(f"Unsupported workbook_type: {wb_type}")
        return 2
    try:
        create_workbook(spec, output_path)
    except Exception as exc:
        print(f"ERROR: {exc}")
        return 3
    print(f"Workbook created: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
